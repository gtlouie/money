'''需求：
    1.获取东方财富每日实时数据，数据来源：http://quote.eastmoney.com/zixuan/?from=data
    2.客户可以自定义股票代码，自定义股票指标
    3.程序运行后，每隔30（时间可自定义）分钟， 记录数据，将数据写入excel中
    4.客户可以自定义文件存放路径 ，生成的Excel 路径

    注*：#号代表注释，程序不运行这句话，客户需要修改的地方是注释下面的语句
'''

#客户可以自定义修改内容区域开始

#自定义股票代码为stockCode,股票名称为：stockName
stockCode='002425'
stockName='凯撒文化'

#自定义文件存放地址,即程序运行后，生成的excel在该路径，如：D:\\金融数据\\
filePath='/Users/gttester/Desktop/PythonWork/NightMoney/moneyFramework/file/'


#自定义时间间隔,即程序运行后，每隔多长时间读取东方财富实时数据，  如：sleepTime=10  10 代表 10 秒钟 ，直接修改数字即可
sleepTime=10

#客户需要自定义的股票指标，客户只需要在对应语句前面添加一个#号即可，如 客户不需要上市时间，   # "f26":"上市日期"
declareData={
    "f12":"代码",
    "f14":"名称",
    "f100":"所属板块",
    "f102":"所属地区板块",
    "f103":"所属概念板块",
    "f2":"最新价",
    "f3":"涨跌幅",
    "f4":"涨跌额",
    "f5":"总手",
    "f30":"现手",
    "f31":"买入价",
    "f32":"卖出价",
    "f18":"昨收",
    "f6":"成交额",
    "f8":"换手率",
    "f7":"振幅",
    "f10":"量比",
    "f22":"涨速",
    "f9":"市盈率",
    "f15":"最高价",
    "f16":"最低价",
    "f17":"开盘价",
    "f62":"主力净流入",
    "f63":"集合竞价",
    "f64":"超大单流入",
    "f65":"超大单流出",
    "f66":"超大单净额",
    "f69":"超大单净占比",
    "f70":"大单流入",
    "f71":"大单流出",
    "f72":"大单净额",
    "f75":"大单净占比",
    "f76":"中单流入",
    "f77":"中单流出",
    "f78":"中单净额",
    "f81":"中单净占比",
    "f82":"小单流入",
    "f83":"小单流出",
    "f84":"小单净额",
    "f87":"小单净占比",
    "f88":"当日DDX",
    "f89":"当日DDY",
    "f90":"当日DDZ",
    "f91":"5日DDX",
    "f92":"5日DDY",
    "f94":"10日DDX",
    "f95":"10日DDY",
    "f97":"DDX飘红天数(连续)",
    "f98":"DDX飘红天数(5日)",
    "f99":"DDX飘红天数(10日)",
    "f38":"总股本",
    "f39":"流通股",
    "f36":"人均持股数",
    "f112":"每股收益",
    "f113":"每股净资产",
    "f37":"净资产收益率(加权)",
    "f40":"营业收入",
    "f41":"营业收入同比",
    "f42":"营业利润",
    "f43":"投资收益",
    "f44":"利润总额",
    "f45":"净利润",
    "f46":"净利润同比",
    "f47":"未分配利润",
    "f48":"每股未分配利润",
    "f49":"毛利率",
    "f50":"总资产",
    "f51":"流动资产",
    "f52":"固定资产",
    "f53":"无形资产",
    "f54":"总负债",
    "f55":"流动负债",
    "f56":"长期负债",
    "f57":"资产负债比率",
    "f58":"股东权益",
    "f59":"股东权益比",
    "f60":"公积金",
    "f61":"每股公积金",
    #"f26":"上市日期"
}



#客户可以自定义修改内容区域结束



#=======================================================================================================================
#下面是程序，不建议修改下面代码，如不小心修改，请使用最初代码进行复制粘贴
#导入程序所需架包
import urllib3
import sseclient
import json
import datetime
import xlsxwriter
import os
import openpyxl
from openpyxl.utils import get_column_letter




class hyNeedData:
    #根据股票id,获取secids
    def get_secids(self,stockCode):
        #如果股票代码第一个数字是0，secids=0.stockCode  ，如果是6，secids=1.stockCode   如果是3 ，secids=0.stockCode
        if stockCode[:1]=='6':
            secids = '1.' + str(stockCode)
        else:
            secids = '0.' + str(stockCode)
        return secids

    #获得客户自定义股票指标，key ,value
    def get_keyVal(self):
        # 获取key数组
        keyArr = []
        # 获取Value数组
        valueArr = []
        for key in declareData:
            keyArr.append(key)
            valueArr.append(declareData[key])
        return keyArr,valueArr

    #获得客户需要查看的股票指标数据
    def get_need_declareData(self):
        # 获取需要查询的字段
        feilds = ""
        for key in declareData:
            feilds = feilds + key + ","
        # 返回需要显示的列
        feilds = feilds[:-1]
        return feilds


    #请求东方财富接口，并获得数据
    def get_targetData(self):
        http = urllib3.PoolManager()
        secids=hyNeedData().get_secids(stockCode)
        feilds=hyNeedData().get_need_declareData()
        url = "https://62.push2.eastmoney.com/api/qt/ulist/sse?invt=3&pi=0&pz=1&mpi=2000&secids=" + secids + "&ut=6d2ffaa6a585d612eda28417681d58fb&fields=" + feilds
        print(url)
        # 获取请求时间
        nowTime = datetime.datetime.now()
        print("数据请求时间"+str(nowTime))
        response=http.request('GET', url, preload_content=False)

        data={}
        client = sseclient.SSEClient(response)
        for msg in client.events():
            if msg.data != None:
                data = json.loads(msg.data)["data"]["diff"]["0"]
                print("请求后返回的数据" + str(data))
            else:
                data = None
            break
        client.close()
        return data,nowTime

    # 根据东方财富的返回的原始数据，进行处理，转成和东方财富页面的数据一样
    def get_handleData(self,data):
        print(data)
        keyArr,valueArr=hyNeedData().get_keyVal()
        # 申明预期插入excel 数组
        excelData=[[0] for i in range(len(keyArr))]
        print(excelData)

        print(1111111)
        for key in data:
            index = keyArr.index(key)
            val = hyNeedData().handData(key, data[key])
            excelData[keyArr.index(key)] = val
        print("处理后的数据"+str(excelData))
        return excelData,valueArr

    #数据写入excel中，生成的数据添加时间点
    def write_excelData(self,valueArr,targetData,nowTime):
        #时间格式化
        nowTimeStr = nowTime.strftime("%Y-%m-%d %H:%M:%S")
        timeStr=nowTime.strftime("%Y-%m-%d")
        valueArr.insert(0,"时间")
        targetData.insert(0,nowTimeStr)
        print(valueArr)
        print(targetData)
        filePathStr =filePath+str(stockName)+"-"+timeStr+'.xlsx'
        worksheetName=stockName+"_"+timeStr+"_数据"
        #判断文件是否存在，不存在就创建
        if os.path.exists(filePathStr):
           print(1)
           # 打开一个将写的文件
           workbook = openpyxl.load_workbook(filePathStr)
           # 在将写的文件创建sheet
           worksheet = workbook[worksheetName]
           worksheet.append(targetData)
           # 保存数据
           workbook.save(filePathStr)

        else:
            print(2)
            # 设置宽度
            columnDimensions = hyNeedData().set_width(len(valueArr))
            # 打开一个将写的文件
            workbook = openpyxl.Workbook()
            # 在将写的文件创建sheetworksheetName
            worksheet = workbook.create_sheet(index=0, title=worksheetName)
            # 设置宽度样式
            for i in range(len(columnDimensions)):
                worksheet.column_dimensions[columnDimensions[i]].width = 20.0

            # 标头写入excel
            for i in range(len(valueArr)):
                worksheet.cell(1, i + 1).value = valueArr[i]
            # 数据写入excel
            for i in range(len(targetData)):
                worksheet.cell(2, i + 1).value = targetData[i]
            # 保存数据
            workbook.save(filePathStr)



    #设置列表列数，方便设置宽度
    def set_width(self,listlLen):

        colArr = []
        for i in range(listlLen):
            letter = get_column_letter(i + 1)
            colArr.append(letter)
        return colArr

    # 处理数据，将数据转为万、亿
    def str_of_num(self,num):
        '''
        递归实现，精确为最大单位值 + 小数点后2位
        '''
        oldnum = num
        num = abs(num)

        def strofsize(num, level):
            if level >= 2:
                return num, level
            elif num >= 10000:
                num /= 10000
                level += 1
                return strofsize(num, level)
            else:
                return num, level

        units = ['', '万', '亿']
        num, level = strofsize(num, 0)
        if level > len(units):
            level -= 1

        if oldnum > 0:
            return '{}{}'.format(round(num, 2), units[level])
        else:
            return '{}{}'.format(round(-num, 2), units[level])

    # 处理数据格式
    def handData(self,key, value):
        val = ''

        if key in ['f2', 'f4', 'f31', 'f32', 'f18', 'f9', 'f15', 'f16', 'f17', 'f90']:
            val = value / 100
        elif key in ['f3', 'f7', 'f8', 'f22', 'f69', 'f75', 'f81', 'f87']:
            val = str(value / 100) + '%'
        elif key in ['f5', 'f6', 'f62', 'f64', 'f65', 'f66', 'f70', 'f71', 'f72', 'f76', 'f77', 'f78', 'f82', 'f83',
                     'f84', 'f38', 'f39', 'f36', 'f40', 'f42', 'f43', 'f44', 'f45', 'f47', 'f50', 'f51', 'f52', 'f53',
                     'f54', 'f55', 'f56', 'f58', 'f60']:
            val = hyNeedData().str_of_num(value)
        elif key in ['f88', 'f89', 'f91', 'f92', 'f93', 'f94', 'f95']:
            val = value / 1000
        elif key in ['f63']:
            val = round(value / 10000)
        elif key in ['f112', 'f113', 'f41', 'f46', 'f48', 'f49', 'f57', 'f59', 'f61']:
            val = round(value, 2)
        elif key == 'f10':
            val = value / 100


        else:
            if value == None or value == '':
                val = "-"
            else:
                val = value
        return val




if __name__=='__main__':
    #程序运行的时间
    data,nowTime=hyNeedData().get_targetData()
    excelData,valueArr=hyNeedData().get_handleData(data)
    print(excelData)
    hyNeedData().write_excelData(valueArr,excelData,nowTime)



